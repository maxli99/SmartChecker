# -*- coding: utf-8 -*-
import re,yaml
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple
from libs.excelchart import ChartFactory,Reference

def dict2list(dictdata,ordered_keys=None):
    """Data is a dict, example:
    _data = {
    'featurecode' : ['5714','5714','5118'],
    'featurename' : ['fea5714','fea5713','fea5118'],
    'capacity'    : [650000,1000000,1000000],
    }
    """
    keys = ordered_keys
    if not keys:
        keys = dictdata.keys()

    _data = [dictdata[key] for key in keys]
    return zip(*_data)

class ExcelReport(object):
    def __init__(self,template_file=None):
        self.template_file = template_file
        self.workbook = None
        
        if isinstance(template_file,str):
            self.load_template(template_file)
            
    
    def load_template(self,template_file):
        self.workbook = load_workbook(template_file)
    
    def fill(self,table):
        """fill the data to worksheet and make the chart
        """
        if not table._data:
            return None

        if not table.sheetname:
            worksheet = self.workbook.active
        else:
            worksheet = self.workbook.get_sheet_by_name(table.sheetname)

        self.fill_data(table,worksheet)
        self.make_chart(table,worksheet)


    def fill_data(self,table,worksheet):
        """Data example:
        _data = [
            ('5714','MME Subscriber LK',650000),
            ('5713','SGSN Attach Capacity',1000000),
        ]
        """        
        data = table._data
        start_row,start_col = coordinate_to_tuple(table.start_cell)
        start_row +=1     
        for ridx,rowdata in enumerate(data):
            for cidx, value in enumerate(rowdata):
                worksheet.cell(row = start_row + ridx,
                               column = start_col + cidx,
                               value = value)

        return worksheet
               
    def make_chart(self,table,worksheet):
        if not hasattr(table,'chart'):
            return None

        factory = ChartFactory()
        chart = factory.create_chart(table.chart)
        data_setting = table.chart['data']
        data = Reference(worksheet,**data_setting)
        xaxis = Reference(worksheet,**table.chart['xaxis'])

        chart.add_data(data)
        chart.chart.set_categories(xaxis)
        chart.add_to_sheet(worksheet)

    def save(self,filename):
        self.workbook.save(filename)
               
      
class Table(object):
    """Class present the data in Excel.
    """
    def __init__(self,tblinfo):
        if isinstance(tblinfo,file):
            tblinfo = yaml.load(tblinfo)
       
        self._data = None
        self.__dict__.update(tblinfo)

        if not tblinfo.get('start_cell',None):
            self.start_cell = 'A1'

    def load_data(self,data):
        self._data = data
        
        if isinstance(data,list):
            self._data = data
        elif isinstance(data,dict):
            self._data = dict2list(data,ordered_keys=self.column_names)
       



